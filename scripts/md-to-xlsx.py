"""Convert a Markdown file's tables into a multi-sheet Excel workbook.

Each `## Heading` becomes a sheet; the first Markdown table under it becomes
that sheet's content. Header row is styled; columns auto-width; long text wraps.

Usage:
    python scripts/md-to-xlsx.py docs/QA-CHECKLIST.md docs/QA-CHECKLIST.xlsx
"""
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def parse_md(md_text):
    """Return list of (section_title, header_cells, rows)."""
    sections = []
    current_title = "Sheet"
    lines = md_text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        h = re.match(r"^##\s+(.*)", line)
        if h:
            current_title = h.group(1).strip()
        # table starts with a row of pipes followed by a |---| separator
        if line.strip().startswith("|") and i + 1 < len(lines) and re.match(
            r"^\s*\|[\s:|-]+\|\s*$", lines[i + 1]
        ):
            header = [c.strip() for c in line.strip().strip("|").split("|")]
            rows = []
            j = i + 2
            while j < len(lines) and lines[j].strip().startswith("|"):
                cells = [c.strip() for c in lines[j].strip().strip("|").split("|")]
                # pad/truncate to header length
                cells = (cells + [""] * len(header))[: len(header)]
                rows.append(cells)
                j += 1
            sections.append((current_title, header, rows))
            i = j
            continue
        i += 1
    return sections


def sheet_name(title, used):
    # Excel sheet names: <=31 chars, no : \ / ? * [ ]
    name = re.sub(r"[:\\/?*\[\]]", "", title)[:31] or "Sheet"
    base, n = name, 1
    while name in used:
        suffix = f" ({n})"
        name = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def build_workbook(sections):
    wb = Workbook()
    wb.remove(wb.active)
    used = set()

    header_fill = PatternFill("solid", fgColor="F97316")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(vertical="top", horizontal="center", wrap_text=True)
    thin = Side(style="thin", color="E6E2D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for title, header, rows in sections:
        ws = wb.create_sheet(sheet_name(title, used))
        ws.append(header)
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border
        for row in rows:
            ws.append(row)
        # style body + widths
        widths = [len(h) for h in header]
        for r in range(2, ws.max_row + 1):
            for idx, cell in enumerate(ws[r]):
                cell.alignment = wrap
                cell.border = border
                text = str(cell.value or "")
                longest = max((len(s) for s in text.split("\n")), default=0)
                widths[idx] = max(widths[idx], longest)
        for idx, w in enumerate(widths):
            # cap width; wrap handles the rest
            ws.column_dimensions[get_column_letter(idx + 1)].width = min(max(w + 2, 10), 55)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"
    return wb


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "docs/QA-CHECKLIST.md")
    out = Path(sys.argv[2] if len(sys.argv) > 2 else src.with_suffix(".xlsx"))
    sections = parse_md(src.read_text(encoding="utf-8"))
    if not sections:
        print("No tables found.")
        sys.exit(1)
    wb = build_workbook(sections)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out} — {len(sections)} sheet(s):")
    for title, header, rows in sections:
        print(f"  - {title}: {len(rows)} rows x {len(header)} cols")


if __name__ == "__main__":
    main()
